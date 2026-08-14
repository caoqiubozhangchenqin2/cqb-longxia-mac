from __future__ import annotations

import argparse
import csv
import json
import math
import os
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from PySide6.QtCharts import QChart, QChartView, QLineSeries, QScatterSeries, QValueAxis
from PySide6.QtCore import QDate, QRect, Qt, Signal
from PySide6.QtGui import QColor, QCursor, QFont, QFontDatabase, QIcon, QPainter, QPdfWriter, QPen, QPixmap
from PySide6.QtWidgets import (
    QApplication, QComboBox, QDateEdit, QDialog, QDialogButtonBox, QDoubleSpinBox,
    QCheckBox, QFileDialog, QFormLayout, QFrame, QGridLayout, QGroupBox, QHBoxLayout, QHeaderView,
    QLabel, QLineEdit, QListWidget, QMainWindow, QMessageBox, QPushButton, QScrollArea,
    QSizePolicy, QStackedWidget, QTableWidget, QTableWidgetItem, QTabWidget, QTextEdit, QToolTip,
    QVBoxLayout, QWidget,
)

from database import HeightDatabase
from standards import age_years, evaluate_height, midparental_target_height, reference_at_age
from theme import APP_STYLE


APP_NAME = "身高小助理"


def configure_chinese_font(app):
    """Load a known Windows Chinese font explicitly, including in packaged/offscreen mode."""
    candidates = [
        Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "msyh.ttc",
        Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "simhei.ttf",
    ]
    family = "Microsoft YaHei UI"
    for font_path in candidates:
        if font_path.exists():
            font_id = QFontDatabase.addApplicationFont(str(font_path))
            families = QFontDatabase.applicationFontFamilies(font_id) if font_id >= 0 else []
            if families:
                family = families[0]
                break
    app.setFont(QFont(family, 10))


def app_root():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def bundled_path(*parts):
    base = Path(getattr(sys, "_MEIPASS", app_root()))
    return base.joinpath(*parts)


def default_data_dir():
    custom = os.environ.get("HEIGHT_ASSISTANT_DATA_DIR")
    return Path(custom) if custom else app_root() / "data"


def item(value, align=Qt.AlignCenter):
    cell = QTableWidgetItem("" if value is None else str(value))
    cell.setTextAlignment(align)
    return cell


class BackgroundWidget(QWidget):
    def __init__(self, image_path=None, parent=None):
        super().__init__(parent)
        self.background = QPixmap(str(image_path)) if image_path and Path(image_path).exists() else QPixmap()

    def paintEvent(self, event):
        painter = QPainter(self)
        if not self.background.isNull():
            scaled = self.background.scaled(self.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            x = (scaled.width() - self.width()) // 2
            y = (scaled.height() - self.height()) // 2
            painter.setOpacity(0.72)
            painter.drawPixmap(QRect(0, 0, self.width(), self.height()), scaled, QRect(x, y, self.width(), self.height()))
        painter.end()
        super().paintEvent(event)


def clear_layout(layout):
    while layout.count():
        child = layout.takeAt(0)
        if child.widget():
            child.widget().deleteLater()


class Card(QFrame):
    def __init__(self, title, value="—", accent=False):
        super().__init__()
        self.setObjectName("Card")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 15, 18, 15)
        title_label = QLabel(title)
        title_label.setObjectName("CardTitle")
        self.value = QLabel(value)
        self.value.setObjectName("AccentValue" if accent else "CardValue")
        layout.addWidget(title_label)
        layout.addWidget(self.value)


class GrowthChart(QChartView):
    TITLES = {
        "height": "身高变化与同龄参考",
        "velocity": "相邻测量区间的年化增长速度",
        "weight": "体重变化趋势",
        "bmi": "BMI 变化趋势",
        "deviation": "身高与同龄中位数的差值",
    }

    def __init__(self, db=None):
        super().__init__()
        self.db = db
        self.setRenderHint(QPainter.Antialiasing)
        self.setMinimumHeight(330)
        self.setMouseTracking(True)
        self.setRubberBand(QChartView.NoRubberBand)
        self.setBackgroundBrush(QColor("transparent"))
        self.last_point_details = []
        self._panning = False
        self._last_pan_pos = None

    def zoom_in(self):
        if self.chart():
            self.chart().zoom(1.25)

    def zoom_out(self):
        if self.chart():
            self.chart().zoom(0.8)

    def reset_zoom(self):
        if self.chart():
            self.chart().zoomReset()

    def wheelEvent(self, event):
        if event.angleDelta().y() > 0:
            self.zoom_in()
        else:
            self.zoom_out()
        event.accept()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.RightButton:
            self.reset_zoom()
            event.accept()
            return
        if event.button() == Qt.LeftButton and self._panning:
            self._panning = False
            self._last_pan_pos = None
            self.unsetCursor()
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._panning = True
            self._last_pan_pos = event.position()
            self.setCursor(Qt.ClosedHandCursor)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._panning and self._last_pan_pos is not None:
            delta = event.position() - self._last_pan_pos
            if self.chart():
                self.chart().scroll(-delta.x(), delta.y())
            self._last_pan_pos = event.position()
            event.accept()
            return
        super().mouseMoveEvent(event)

    def _show_tooltip(self, point, state, lookup):
        if not state:
            QToolTip.hideText()
            return
        key = (round(point.x(), 5), round(point.y(), 5))
        text = lookup.get(key, f"年龄：{point.x():.2f} 岁\n数值：{point.y():.2f}")
        QToolTip.showText(QCursor.pos(), text, self)

    def _add_interactive_series(self, chart, entries, name, color):
        line = QLineSeries(); line.setName(name); line.setPen(QPen(QColor(color), 3))
        dots = QScatterSeries(); dots.setName(f"{name}数据点"); dots.setMarkerSize(10)
        dots.setColor(QColor(color)); dots.setBorderColor(QColor("#FFFFFF"))
        note_dots = QScatterSeries(); note_dots.setName("含备注"); note_dots.setMarkerSize(15)
        note_dots.setColor(QColor("#FF765E")); note_dots.setBorderColor(QColor("#FFFFFF"))
        lookup = {}; note_lookup = {}
        for entry in entries:
            x, y, details = entry[:3]
            has_note = bool(entry[3]) if len(entry) > 3 else False
            line.append(x, y); dots.append(x, y)
            lookup[(round(x, 5), round(y, 5))] = details
            if has_note:
                note_dots.append(x, y)
                note_lookup[(round(x, 5), round(y, 5))] = details
        dots.hovered.connect(lambda point, state, values=lookup: self._show_tooltip(point, state, values))
        chart.addSeries(line); chart.addSeries(dots)
        markers = chart.legend().markers(dots)
        if markers:
            markers[0].setVisible(False)
        if note_lookup:
            note_dots.hovered.connect(lambda point, state, values=note_lookup: self._show_tooltip(point, state, values))
            chart.addSeries(note_dots)
        return [(entry[0], entry[1]) for entry in entries]

    def _measurement_details(self, child, row):
        age = age_years(child["birth_date"], row["measured_at"])
        lines = [f'<b>{row["measured_at"]}</b>', f'年龄：{age:.2f} 岁', f'身高：{row["height_cm"]:.1f} cm']
        if row["weight_jin"] is not None:
            lines.append(f'体重：{row["weight_jin"]:.1f} 斤（{row["weight_jin"] / 2:.1f} kg）')
        if row.get("method"):
            lines.append(f'方式：{row["method"]}')
        if self.db:
            meds = self.db.measurement_medications(row["id"])
            if meds:
                lines.append("当日用药：" + "、".join(med["snapshot_name"] for med in meds))
        if row.get("notes"):
            lines.append(f'<b style="color:#FFB09F">备注：{row["notes"]}</b>')
        return "<br>".join(lines)

    def render_data(self, child, rows, mode="height", show_reference=True):
        chart = QChart()
        chart.setAnimationOptions(QChart.NoAnimation)
        chart.setBackgroundBrush(QColor("#FFFFFF"))
        chart.setPlotAreaBackgroundVisible(True)
        chart.setPlotAreaBackgroundBrush(QColor("#FBFDFD"))
        chart.legend().setVisible(True); chart.legend().setAlignment(Qt.AlignBottom)
        chart.setTitle(self.TITLES.get(mode, "成长趋势"))
        axis_x = QValueAxis(); axis_x.setTitleText("年龄（岁）"); axis_x.setLabelFormat("%.1f"); axis_x.setGridLineColor(QColor("#E4ECEF"))
        axis_y = QValueAxis(); axis_y.setGridLineColor(QColor("#E4ECEF"))
        points = []; entries = []

        if mode == "height":
            for row in rows:
                x = age_years(child["birth_date"], row["measured_at"]); y = row["height_cm"]
                entries.append((x, y, self._measurement_details(child, row), bool(row.get("notes"))))
            points.extend(self._add_interactive_series(chart, entries, "实际身高", "#16877F"))
            actual_ages = [entry[0] for entry in entries]
            if show_reference and actual_ages:
                start_age = max(7.0, min(actual_ages) - 0.5); end_age = min(18.0, max(actual_ages) + 0.5)
                reference_lines = (
                    ("参考 -2SD", "#E6B663", 0),
                    ("参考 -1SD", "#9DBE72", 1),
                    ("参考中位数", "#72A8C3", 2),
                    ("参考 +1SD", "#B58CC5", 3),
                    ("参考 +2SD", "#D98B88", 4),
                )
                for name, color, value_index in reference_lines:
                    series = QLineSeries(); series.setName(name); series.setPen(QPen(QColor(color), 1.5, Qt.DashLine))
                    for quarter in range(int(start_age * 4), int(end_age * 4) + 2):
                        age = quarter / 4; ref = reference_at_age(child["gender"], age)
                        if ref:
                            value = ref[value_index]; series.append(age, value); points.append((age, value))
                    chart.addSeries(series)
            axis_y.setTitleText("身高（cm）")
        elif mode == "velocity":
            for before, after in zip(rows, rows[1:]):
                days = (date.fromisoformat(after["measured_at"]) - date.fromisoformat(before["measured_at"])).days
                if days <= 0: continue
                delta = after["height_cm"] - before["height_cm"]
                x = age_years(child["birth_date"], after["measured_at"]); y = delta * 365.2425 / days
                details = f'<b>{before["measured_at"]} → {after["measured_at"]}</b><br>间隔：{days} 天<br>长高：{delta:.1f} cm<br>年化速度：{y:.2f} cm/年'
                note_parts=[]
                if before.get("notes"):note_parts.append(f'起点：{before["notes"]}')
                if after.get("notes"):note_parts.append(f'终点：{after["notes"]}')
                if note_parts:details += '<br><b style="color:#FFB09F">备注：' + '；'.join(note_parts) + '</b>'
                entries.append((x, y, details, bool(note_parts)))
            points.extend(self._add_interactive_series(chart, entries, "年化增长", "#8A67B8")); axis_y.setTitleText("cm / 年")
        elif mode == "weight":
            for row in rows:
                if row["weight_jin"] is None: continue
                x = age_years(child["birth_date"], row["measured_at"]); y = row["weight_jin"]
                entries.append((x, y, self._measurement_details(child, row), bool(row.get("notes"))))
            points.extend(self._add_interactive_series(chart, entries, "体重", "#E29B55")); axis_y.setTitleText("体重（斤）")
        elif mode == "bmi":
            for row in rows:
                if row["weight_jin"] is None or row["height_cm"] <= 0: continue
                x = age_years(child["birth_date"], row["measured_at"])
                y = (row["weight_jin"] / 2) / math.pow(row["height_cm"] / 100, 2)
                details = self._measurement_details(child, row) + f"<br><b>BMI：{y:.2f}</b>"
                entries.append((x, y, details, bool(row.get("notes"))))
            points.extend(self._add_interactive_series(chart, entries, "BMI", "#D76D8C")); axis_y.setTitleText("BMI（kg/m²）")
        else:
            for row in rows:
                x = age_years(child["birth_date"], row["measured_at"]); ref = reference_at_age(child["gender"], x)
                if not ref: continue
                y = row["height_cm"] - ref[2]
                details = self._measurement_details(child, row) + f"<br><b>较中位数：{y:+.2f} cm</b>"
                entries.append((x, y, details, bool(row.get("notes"))))
            points.extend(self._add_interactive_series(chart, entries, "距同龄中位数", "#4D91B8")); axis_y.setTitleText("差值（cm）")
            zero = QLineSeries(); zero.setName("同龄中位数基线")
            if entries:
                zero.append(min(entry[0] for entry in entries), 0); zero.append(max(entry[0] for entry in entries), 0)
            zero.setPen(QPen(QColor("#9EADB2"), 1.5, Qt.DashLine)); chart.addSeries(zero)

        chart.addAxis(axis_x, Qt.AlignBottom); chart.addAxis(axis_y, Qt.AlignLeft)
        for series in chart.series(): series.attachAxis(axis_x); series.attachAxis(axis_y)
        if points:
            xs, ys = zip(*points); xpad = max(0.20, (max(xs)-min(xs))*.06); ypad = max(1.0, (max(ys)-min(ys))*.10)
            axis_x.setRange(max(0, min(xs)-xpad), max(xs)+xpad); axis_y.setRange(min(ys)-ypad, max(ys)+ypad)
        else:
            axis_x.setRange(0, 18); axis_y.setRange(0, 180)
        self.last_point_details = entries
        self.setChart(chart)


class HomePage(QWidget):
    def __init__(self, db):
        super().__init__(); self.db = db
        root = QVBoxLayout(self); root.setSpacing(14)
        cards = QHBoxLayout(); self.latest = Card("最近身高", accent=True); self.growth = Card("累计增长")
        self.count = Card("记录次数"); self.status = Card("参考位置")
        for card in (self.latest, self.growth, self.count, self.status): cards.addWidget(card)
        root.addLayout(cards)
        self.chart = GrowthChart(db); root.addWidget(self.chart, 1)
        self.chart_help=QLabel("图表操作：悬停圆点查看数据与备注｜鼠标滚轮缩放｜按住左键拖动图表｜右键恢复全图");self.chart_help.setObjectName("ChartHelp");root.addWidget(self.chart_help)
        self.table = QTableWidget(0, 4); self.table.setHorizontalHeaderLabels(["日期", "身高", "体重", "备注"])
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setMaximumHeight(210); self.table.setAlternatingRowColors(True)
        root.addWidget(self.table)

    def refresh(self, child):
        if not child: return
        rows = self.db.measurements(child["id"])
        summary = self.db.summary(child["id"])
        latest = summary["latest"]
        self.latest.value.setText(f'{latest["height_cm"]:.1f} cm' if latest else "暂无")
        self.growth.value.setText(f'+{summary["growth"]:.1f} cm' if summary["growth"] is not None else "暂无")
        self.count.value.setText(f'{summary["count"]} 次')
        if latest:
            age = age_years(child["birth_date"], latest["measured_at"])
            label, _ = evaluate_height(child["gender"], age, latest["height_cm"])
            self.status.value.setText(label)
        else: self.status.value.setText("暂无")
        self.chart.render_data(child, rows)
        recent = list(reversed(rows[-5:])); self.table.setRowCount(len(recent))
        for r, row in enumerate(recent):
            vals = [row["measured_at"], f'{row["height_cm"]:.1f} cm',
                    "—" if row["weight_jin"] is None else f'{row["weight_jin"]:.1f} 斤', row["notes"]]
            for c, val in enumerate(vals): self.table.setItem(r, c, item(val, Qt.AlignLeft | Qt.AlignVCenter if c == 3 else Qt.AlignCenter))


class AddPage(QWidget):
    saved = Signal()
    def __init__(self, db, current_child):
        super().__init__(); self.db = db; self.current_child = current_child
        root = QVBoxLayout(self)
        form_box = QGroupBox("新增测量记录"); form = QGridLayout(form_box)
        self.measured = QDateEdit(QDate.currentDate()); self.measured.setCalendarPopup(True); self.measured.setDisplayFormat("yyyy-MM-dd")
        self.height = QDoubleSpinBox(); self.height.setRange(30, 250); self.height.setDecimals(1); self.height.setSuffix(" cm")
        self.weight = QDoubleSpinBox(); self.weight.setRange(0, 400); self.weight.setDecimals(1); self.weight.setSuffix(" 斤"); self.weight.setSpecialValueText("未填写")
        self.method = QComboBox(); self.method.addItems(["", "家中测量", "医院测量", "学校测量", "其他"])
        self.notes = QTextEdit(); self.notes.setMaximumHeight(85); self.notes.setPlaceholderText("睡眠、运动、复查情况或其他需要保留的信息")
        form.addWidget(QLabel("测量日期"),0,0); form.addWidget(self.measured,0,1)
        form.addWidget(QLabel("身高"),0,2); form.addWidget(self.height,0,3)
        form.addWidget(QLabel("体重"),1,0); form.addWidget(self.weight,1,1)
        form.addWidget(QLabel("测量方式"),1,2); form.addWidget(self.method,1,3)
        form.addWidget(QLabel("备注"),2,0); form.addWidget(self.notes,2,1,1,3)
        root.addWidget(form_box)

        med_box = QGroupBox("当日用药（可同时录入多种）"); med_layout = QVBoxLayout(med_box)
        self.med_table = QTableWidget(0, 5); self.med_table.setHorizontalHeaderLabels(["药品名称", "剂量", "单位", "频次", "状态"])
        self.med_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        med_layout.addWidget(self.med_table)
        row = QHBoxLayout(); add = QPushButton("＋ 添加一种药"); add.clicked.connect(self.add_med_row)
        load = QPushButton("载入当日使用中的药品"); load.clicked.connect(self.load_active_meds)
        row.addWidget(add); row.addWidget(load); row.addStretch(); med_layout.addLayout(row)
        root.addWidget(med_box)
        save = QPushButton("保存本次记录"); save.setObjectName("Primary"); save.clicked.connect(self.save); root.addWidget(save, alignment=Qt.AlignRight)
        root.addStretch()

    def add_med_row(self, values=None):
        r = self.med_table.rowCount(); self.med_table.insertRow(r); values = values or {}
        for c, key in enumerate(("name", "dose", "unit", "frequency", "status")):
            self.med_table.setItem(r, c, item(values.get(key, ""), Qt.AlignLeft | Qt.AlignVCenter))

    def load_active_meds(self):
        child = self.current_child()
        if not child: return
        self.med_table.setRowCount(0)
        for med in self.db.active_medications(child["id"], self.measured.date().toString("yyyy-MM-dd")):
            self.add_med_row({"name": med["medication_name"], "dose": med["dose"] or "", "unit": med["unit"],
                              "frequency": med["frequency"], "status": med["status"]})

    def save(self):
        child = self.current_child()
        if not child: QMessageBox.warning(self, APP_NAME, "请先在设置中添加孩子。"); return
        if self.height.value() <= 30: QMessageBox.warning(self, APP_NAME, "请填写有效身高。"); return
        snapshots = []
        for r in range(self.med_table.rowCount()):
            def txt(c): return self.med_table.item(r,c).text().strip() if self.med_table.item(r,c) else ""
            name = txt(0)
            if not name: continue
            med_id = self.db.get_or_create_medication(child["id"], name, default_unit=txt(2))
            try: dose = float(txt(1)) if txt(1) else None
            except ValueError: dose = None
            snapshots.append({"medication_id": med_id, "name": name, "dose": dose, "unit": txt(2), "frequency": txt(3), "status": txt(4)})
        weight = self.weight.value() or None
        self.db.add_measurement(child["id"], self.measured.date().toString("yyyy-MM-dd"), self.height.value(),
                                weight, self.method.currentText(), self.notes.toPlainText().strip(), medication_snapshots=snapshots)
        QMessageBox.information(self, APP_NAME, "记录已安全保存。")
        self.notes.clear(); self.med_table.setRowCount(0); self.saved.emit()


class EditMeasurementDialog(QDialog):
    def __init__(self, record, parent=None):
        super().__init__(parent);self.record=record;self.setWindowTitle("编辑历史记录");self.resize(620,520)
        root=QVBoxLayout(self);title=QLabel("编辑测量数据");title.setObjectName("PageTitle");root.addWidget(title)
        hint=QLabel("备注使用大文本框完整显示。点击“撤回 / 取消”不会保存任何变化。");hint.setObjectName("Muted");root.addWidget(hint)
        form_box=QGroupBox("测量内容");form=QFormLayout(form_box)
        self.measured=QDateEdit(QDate.fromString(record["measured_at"],"yyyy-MM-dd"));self.measured.setCalendarPopup(True);self.measured.setDisplayFormat("yyyy-MM-dd")
        self.height=QDoubleSpinBox();self.height.setRange(30,250);self.height.setDecimals(1);self.height.setSuffix(" cm");self.height.setValue(record["height_cm"])
        self.weight=QDoubleSpinBox();self.weight.setRange(0,400);self.weight.setDecimals(1);self.weight.setSuffix(" 斤");self.weight.setSpecialValueText("未填写");self.weight.setValue(record["weight_jin"] or 0)
        self.method=QComboBox();self.method.setEditable(True);self.method.addItems(["","家中测量","医院测量","学校测量","其他"]);self.method.setCurrentText(record.get("method", ""))
        self.recheck=QCheckBox("这条记录需要复测或再次确认");self.recheck.setChecked(bool(record.get("needs_recheck")))
        self.notes=QTextEdit();self.notes.setPlainText(record.get("notes", ""));self.notes.setMinimumHeight(190);self.notes.setPlaceholderText("在这里完整填写复查、用药、测量环境等备注")
        for label,widget in (("测量日期",self.measured),("身高",self.height),("体重",self.weight),("测量方式",self.method),("记录状态",self.recheck),("备注",self.notes)):form.addRow(label,widget)
        root.addWidget(form_box,1)
        buttons=QHBoxLayout();cancel=QPushButton("撤回 / 取消");cancel.setObjectName("Undo");cancel.clicked.connect(self.reject);save=QPushButton("保存修改");save.setObjectName("Primary");save.clicked.connect(self.accept);buttons.addStretch();buttons.addWidget(cancel);buttons.addWidget(save);root.addLayout(buttons)

    def values(self):
        return {"measured_at":self.measured.date().toString("yyyy-MM-dd"),"height_cm":self.height.value(),"weight_jin":self.weight.value() or None,"method":self.method.currentText().strip(),"notes":self.notes.toPlainText().strip(),"needs_recheck":self.recheck.isChecked()}


class HistoryPage(QWidget):
    changed = Signal()
    def __init__(self, db, current_child):
        super().__init__(); self.db=db; self.current_child=current_child
        root=QVBoxLayout(self); actions=QHBoxLayout();
        actions.addWidget(QLabel("日期顺序")); self.sort_order=QComboBox(); self.sort_order.addItem("最新在前",Qt.DescendingOrder); self.sort_order.addItem("最早在前",Qt.AscendingOrder); self.sort_order.currentIndexChanged.connect(self.apply_sort); actions.addWidget(self.sort_order)
        self.search=QLineEdit(); self.search.setPlaceholderText("搜索日期、方式或备注"); self.search.setMaximumWidth(240); self.search.textChanged.connect(self.apply_filter); actions.addWidget(self.search)
        for label, func in (("导出 Excel", self.export_excel),("导出 CSV", self.export_csv),("导出 JSON", self.export_json)):
            b=QPushButton(label); b.clicked.connect(func); actions.addWidget(b)
        actions.addStretch()
        root.addLayout(actions)
        edit_actions=QHBoxLayout();hint=QLabel("双击一行也可以打开大窗口编辑");hint.setObjectName("Muted");edit_actions.addWidget(hint);edit_actions.addStretch()
        edit=QPushButton("编辑选中记录");edit.setObjectName("Primary");edit.clicked.connect(self.edit_selected);edit_actions.addWidget(edit)
        self.undo_button=QPushButton("撤回上次修改");self.undo_button.setObjectName("Undo");self.undo_button.setEnabled(False);self.undo_button.clicked.connect(self.undo_last_edit);edit_actions.addWidget(self.undo_button)
        delete=QPushButton("归档选中记录");delete.setObjectName("Danger");delete.clicked.connect(self.delete_selected);edit_actions.addWidget(delete);root.addLayout(edit_actions)
        self.table=QTableWidget(0,6); self.table.setHorizontalHeaderLabels(["日期","年龄","身高(cm)","体重(斤)","测量方式","备注"])
        self.table.horizontalHeader().setSectionResizeMode(5,QHeaderView.Stretch); self.table.setAlternatingRowColors(True); self.table.setSelectionBehavior(QTableWidget.SelectRows); self.table.setEditTriggers(QTableWidget.NoEditTriggers);self.table.setSortingEnabled(True);self.table.doubleClicked.connect(self.edit_selected)
        root.addWidget(self.table)
        self.last_edit=None

    def refresh(self, child):
        self.table.setSortingEnabled(False); self.table.setRowCount(0)
        if not child:self.table.setSortingEnabled(True);return
        rows=self.db.measurements(child["id"]); self.table.setRowCount(len(rows))
        for r,row in enumerate(rows):
            age=age_years(child["birth_date"],row["measured_at"])
            vals=(row["measured_at"],f"{age:.2f}",f'{row["height_cm"]:.1f}',"" if row["weight_jin"] is None else f'{row["weight_jin"]:.1f}',row["method"],row["notes"])
            for c,val in enumerate(vals):self.table.setItem(r,c,item(val,Qt.AlignLeft|Qt.AlignVCenter if c==5 else Qt.AlignCenter))
            self.table.item(r,0).setData(Qt.UserRole,row["id"])
        self.table.setSortingEnabled(True); self.apply_sort(); self.apply_filter()

    def apply_sort(self):
        if hasattr(self,"table"):
            self.table.sortItems(0,self.sort_order.currentData())

    def apply_filter(self):
        needle=self.search.text().strip().lower() if hasattr(self,"search") else ""
        for row in range(self.table.rowCount()):
            haystack=" ".join(self.table.item(row,col).text() for col in range(self.table.columnCount()) if self.table.item(row,col)).lower()
            self.table.setRowHidden(row,bool(needle and needle not in haystack))

    def _save_path(self,suffix):
        child=self.current_child(); name=child["name"] if child else "身高记录"
        return QFileDialog.getSaveFileName(self,"导出",str(self.db.export_dir/f"{name}-身高记录.{suffix}"),f"*.{suffix}")[0]

    def export_excel(self):
        path=self._save_path("xlsx"); child=self.current_child()
        if not path or not child:return
        wb=Workbook(); ws=wb.active; ws.title="身高记录"; ws.append(["日期","年龄(岁)","身高(cm)","体重(斤)","测量方式","备注"])
        for row in self.db.measurements(child["id"]): ws.append([row["measured_at"],round(age_years(child["birth_date"],row["measured_at"]),2),row["height_cm"],row["weight_jin"],row["method"],row["notes"]])
        wb.save(path); QMessageBox.information(self,APP_NAME,f"已导出到：\n{path}")

    def export_csv(self):
        path=self._save_path("csv"); child=self.current_child()
        if not path or not child:return
        with open(path,"w",newline="",encoding="utf-8-sig") as f:
            w=csv.writer(f); w.writerow(["日期","身高(cm)","体重(斤)","测量方式","备注"])
            for row in self.db.measurements(child["id"]): w.writerow([row["measured_at"],row["height_cm"],row["weight_jin"],row["method"],row["notes"]])
        QMessageBox.information(self,APP_NAME,f"已导出到：\n{path}")

    def export_json(self):
        path=self._save_path("json"); child=self.current_child()
        if not path or not child:return
        Path(path).write_text(json.dumps({"child":child,"measurements":self.db.measurements(child["id"]),"medication_periods":self.db.medication_periods(child["id"])},ensure_ascii=False,indent=2),encoding="utf-8")
        QMessageBox.information(self,APP_NAME,f"已导出到：\n{path}")

    def selected_measurement_id(self):
        row=self.table.currentRow()
        return self.table.item(row,0).data(Qt.UserRole) if row>=0 and self.table.item(row,0) else None

    def edit_selected(self,*_):
        measurement_id=self.selected_measurement_id()
        if not measurement_id:QMessageBox.information(self,APP_NAME,"请先选择一条历史记录。");return
        record=self.db.measurement(measurement_id)
        if not record:return
        dialog=EditMeasurementDialog(record,self)
        if dialog.exec()!=QDialog.Accepted:return
        try:self.last_edit=self.db.update_measurement(measurement_id,dialog.values())
        except ValueError as exc:QMessageBox.warning(self,APP_NAME,str(exc));return
        self.undo_button.setEnabled(True);self.undo_button.setToolTip(f'恢复 {record["measured_at"]} 保存前的内容');self.changed.emit();QMessageBox.information(self,APP_NAME,"修改已保存，首页和全部图表已刷新。")

    def undo_last_edit(self):
        if not self.last_edit:return
        if QMessageBox.question(self,APP_NAME,f'确认撤回对 {self.last_edit["measured_at"]} 记录的上次修改？')!=QMessageBox.Yes:return
        try:self.db.restore_measurement(self.last_edit)
        except ValueError as exc:QMessageBox.warning(self,APP_NAME,str(exc));return
        self.last_edit=None;self.undo_button.setEnabled(False);self.changed.emit();QMessageBox.information(self,APP_NAME,"上次修改已撤回，首页和全部图表已恢复。")

    def delete_selected(self):
        r=self.table.currentRow()
        if r<0:return
        if QMessageBox.question(self,APP_NAME,"确认归档选中的测量记录？原始数据仍会保留在数据库中。")!=QMessageBox.Yes:return
        self.db.delete_measurement(self.table.item(r,0).data(Qt.UserRole)); self.changed.emit()


class ChartsPage(QWidget):
    def __init__(self,db,current_child):
        super().__init__(); self.db=db; self.current_child=current_child
        root=QVBoxLayout(self); bar=QHBoxLayout(); bar.addWidget(QLabel("图表类型")); self.mode=QComboBox();
        for label,value in (("身高参考","height"),("年化增速","velocity"),("体重","weight"),("BMI","bmi"),("同龄中位差","deviation")):self.mode.addItem(label,value)
        self.mode.currentIndexChanged.connect(self.refresh_current); bar.addWidget(self.mode)
        bar.addWidget(QLabel("观察范围"));self.range=QComboBox();
        for label,days in (("全部",0),("近 6 个月",183),("近 1 年",366),("近 2 年",731),("近 3 年",1096)):self.range.addItem(label,days)
        self.range.currentIndexChanged.connect(self.refresh_current);bar.addWidget(self.range)
        self.show_reference=QCheckBox("显示参考线");self.show_reference.setChecked(True);self.show_reference.toggled.connect(self.refresh_current);bar.addWidget(self.show_reference)
        bar.addStretch()
        png=QPushButton("导出 PNG"); png.clicked.connect(self.export_png); pdf=QPushButton("导出 PDF"); pdf.clicked.connect(self.export_pdf); detail=QPushButton("导出分析表");detail.clicked.connect(self.export_analysis);bar.addWidget(png);bar.addWidget(pdf);bar.addWidget(detail);root.addLayout(bar)
        cards=QHBoxLayout();self.range_count=Card("范围内记录");self.latest_value=Card("最新数值",accent=True);self.period_change=Card("阶段变化");self.avg_velocity=Card("平均年增速");self.target_height_card=Card("成年遗传靶身高")
        for card in (self.range_count,self.latest_value,self.period_change,self.avg_velocity,self.target_height_card):cards.addWidget(card)
        root.addLayout(cards)
        self.target_height_info=QLabel();self.target_height_info.setObjectName("TargetHelp");self.target_height_info.setWordWrap(True);root.addWidget(self.target_height_info)
        self.chart=GrowthChart(db);root.addWidget(self.chart,1)
        self.chart_help=QLabel("图表操作：悬停圆点查看数据与备注｜鼠标滚轮缩放｜按住左键拖动图表｜右键恢复全图");self.chart_help.setObjectName("ChartHelp");root.addWidget(self.chart_help)
        tabs=QTabWidget();
        self.intervals=QTableWidget(0,6);self.intervals.setHorizontalHeaderLabels(["起始日期","结束日期","间隔天数","长高(cm)","年化(cm/年)","体重变化(斤)"]);self.intervals.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch);self.intervals.setAlternatingRowColors(True)
        self.periods=QTableWidget(0,6);self.periods.setHorizontalHeaderLabels(["药品","开始","结束","状态","剂量","说明"]);self.periods.horizontalHeader().setSectionResizeMode(5,QHeaderView.Stretch)
        tabs.addTab(self.intervals,"分段增长明细");tabs.addTab(self.periods,"用药阶段对照");tabs.setMaximumHeight(220);root.addWidget(tabs)
        self.child=None

    def refresh(self,child):
        self.child=child; self.refresh_current()

    def refresh_current(self):
        if not self.child:return
        all_rows=self.db.measurements(self.child["id"]);rows=self.filtered_rows(all_rows)
        target_values=midparental_target_height(self.child["gender"],self.child.get("father_height"),self.child.get("mother_height"))
        if target_values:
            target,target_low,target_high=target_values
            self.target_height_card.value.setText(f"{target:.1f} cm")
            self.target_height_card.setToolTip(f"常用参考范围：{target_low:.1f}–{target_high:.1f} cm")
            self.target_height_info.setText(
                f"遗传身高参考范围：{target_low:.1f}–{target_high:.1f} cm。"
                "此数据独立显示，不参与图表坐标缩放；它是基于父母身高的粗略成年参考，并非确定预测。"
            )
        else:
            self.target_height_card.value.setText("待填写")
            self.target_height_card.setToolTip("请先填写父亲和母亲身高")
            self.target_height_info.setText("遗传靶身高尚未计算：请在“设置 → 孩子与需求”中填写父亲和母亲身高。")
        self.chart.render_data(self.child,rows,self.mode.currentData(),self.show_reference.isChecked())
        self.fill_summary(rows);self.fill_intervals(rows)
        rows=self.db.medication_periods(self.child["id"]);self.periods.setRowCount(len(rows))
        for r,row in enumerate(rows):
            dose="" if row["dose"] is None else f'{row["dose"]} {row["unit"]}'
            vals=(row["medication_name"],row["start_date"],row["end_date"] or "至今",row["status"],dose,row["notes"])
            for c,val in enumerate(vals):self.periods.setItem(r,c,item(val,Qt.AlignLeft|Qt.AlignVCenter if c==5 else Qt.AlignCenter))

    def filtered_rows(self,rows):
        days=self.range.currentData()
        if not days or not rows:return rows
        cutoff=date.fromisoformat(rows[-1]["measured_at"])-timedelta(days=days)
        return [row for row in rows if date.fromisoformat(row["measured_at"])>=cutoff]

    def fill_summary(self,rows):
        self.range_count.value.setText(f"{len(rows)} 次")
        if not rows:
            for card in (self.latest_value,self.period_change,self.avg_velocity):card.value.setText("暂无")
            return
        mode=self.mode.currentData();latest=rows[-1]
        if mode=="weight":value="暂无" if latest["weight_jin"] is None else f'{latest["weight_jin"]:.1f} 斤'
        elif mode=="bmi" and latest["weight_jin"] is not None:value=f'{(latest["weight_jin"]/2)/math.pow(latest["height_cm"]/100,2):.2f}'
        else:value=f'{latest["height_cm"]:.1f} cm'
        self.latest_value.value.setText(value)
        if len(rows)<2:
            self.period_change.value.setText("暂无");self.avg_velocity.value.setText("暂无");return
        delta=rows[-1]["height_cm"]-rows[0]["height_cm"]
        days=(date.fromisoformat(rows[-1]["measured_at"])-date.fromisoformat(rows[0]["measured_at"])).days
        self.period_change.value.setText(f"+{delta:.1f} cm");self.avg_velocity.value.setText(f"{delta*365.2425/days:.2f} cm/年" if days>0 else "暂无")

    def fill_intervals(self,rows):
        pairs=[]
        for before,after in zip(rows,rows[1:]):
            days=(date.fromisoformat(after["measured_at"])-date.fromisoformat(before["measured_at"])).days
            if days<=0:continue
            delta=after["height_cm"]-before["height_cm"];velocity=delta*365.2425/days
            weight_delta=""
            if before["weight_jin"] is not None and after["weight_jin"] is not None:weight_delta=f'{after["weight_jin"]-before["weight_jin"]:+.1f}'
            pairs.append((before["measured_at"],after["measured_at"],days,f'{delta:+.1f}',f'{velocity:.2f}',weight_delta))
        self.intervals.setRowCount(len(pairs))
        for r,values in enumerate(reversed(pairs)):
            for c,value in enumerate(values):self.intervals.setItem(r,c,item(value))

    def export_png(self):
        if not self.child:return
        default=self.db.export_dir/f'{self.child["name"]}-{self.mode.currentText()}.png';path=QFileDialog.getSaveFileName(self,"导出图表",str(default),"PNG 图片 (*.png)")[0]
        if path:self.chart.grab().save(path,"PNG");QMessageBox.information(self,APP_NAME,f"图表已导出：\n{path}")

    def export_pdf(self):
        if not self.child:return
        default=self.db.export_dir/f'{self.child["name"]}-{self.mode.currentText()}.pdf';path=QFileDialog.getSaveFileName(self,"导出图表",str(default),"PDF 文件 (*.pdf)")[0]
        if not path:return
        writer=QPdfWriter(path);writer.setResolution(144);painter=QPainter(writer);self.chart.render(painter);painter.end();QMessageBox.information(self,APP_NAME,f"图表已导出：\n{path}")

    def export_analysis(self):
        if not self.child:return
        path=QFileDialog.getSaveFileName(self,"导出详细分析",str(self.db.export_dir/f'{self.child["name"]}-成长分析.xlsx'),"Excel (*.xlsx)")[0]
        if not path:return
        rows=self.filtered_rows(self.db.measurements(self.child["id"]));wb=Workbook();ws=wb.active;ws.title="测量明细";ws.append(["日期","年龄(岁)","身高(cm)","体重(斤)","BMI","测量方式","备注"])
        for row in rows:
            bmi=(row["weight_jin"]/2)/math.pow(row["height_cm"]/100,2) if row["weight_jin"] is not None else None
            ws.append([row["measured_at"],round(age_years(self.child["birth_date"],row["measured_at"]),3),row["height_cm"],row["weight_jin"],round(bmi,2) if bmi else None,row["method"],row["notes"]])
        interval=wb.create_sheet("分段增长");interval.append(["起始日期","结束日期","间隔天数","长高(cm)","年化(cm/年)"])
        for before,after in zip(rows,rows[1:]):
            days=(date.fromisoformat(after["measured_at"])-date.fromisoformat(before["measured_at"])).days
            if days>0:
                delta=after["height_cm"]-before["height_cm"];interval.append([before["measured_at"],after["measured_at"],days,round(delta,2),round(delta*365.2425/days,2)])
        overview=wb.create_sheet("遗传身高参考")
        overview.append(["项目","数值","说明"])
        overview.append(["父亲身高",self.child.get("father_height"),"cm"])
        overview.append(["母亲身高",self.child.get("mother_height"),"cm"])
        target_values=midparental_target_height(self.child["gender"],self.child.get("father_height"),self.child.get("mother_height"))
        if target_values:
            target,target_low,target_high=target_values
            overview.append(["成年遗传靶身高",target,"cm；父母中位身高公式"])
            overview.append(["常用参考范围",f"{target_low:.1f}–{target_high:.1f}","cm；靶身高 ±10.2 cm"])
        else:
            overview.append(["成年遗传靶身高","无法计算","需要同时填写父亲和母亲身高"])
        overview.append(["重要说明","仅供成长观察参考","不是个体成年身高的确定预测，不能替代骨龄和医生评估"])
        wb.save(path);QMessageBox.information(self,APP_NAME,f"详细分析已导出：\n{path}")


class MedicationDialog(QDialog):
    def __init__(self,parent=None):
        super().__init__(parent);self.setWindowTitle("添加用药阶段");self.resize(480,430);form=QFormLayout(self)
        self.name=QLineEdit();self.name.setPlaceholderText("可输入任意药品名称")
        self.start=QDateEdit(QDate.currentDate());self.start.setCalendarPopup(True);self.start.setDisplayFormat("yyyy-MM-dd")
        self.end=QDateEdit(QDate.currentDate());self.end.setCalendarPopup(True);self.end.setDisplayFormat("yyyy-MM-dd");self.end.setSpecialValueText("未结束");self.end.setMinimumDate(QDate(1900,1,1));self.end.setDate(self.end.minimumDate())
        self.status=QComboBox();self.status.addItems(["使用中","已结束","已停用","待确认"])
        self.dose=QDoubleSpinBox();self.dose.setRange(0,100000);self.dose.setDecimals(3);self.dose.setSpecialValueText("待确认")
        self.unit=QLineEdit();self.unit.setPlaceholderText("例如 mg、IU、支")
        self.frequency=QLineEdit();self.frequency.setPlaceholderText("例如 每晚一次")
        self.route=QLineEdit();self.route.setPlaceholderText("例如 皮下注射、口服")
        self.doctor=QLineEdit();self.notes=QTextEdit();self.notes.setMaximumHeight(70)
        for label,w in (("药品名称",self.name),("开始日期",self.start),("结束日期",self.end),("状态",self.status),("剂量",self.dose),("单位",self.unit),("频次",self.frequency),("用法",self.route),("医生/机构",self.doctor),("说明",self.notes)):form.addRow(label,w)
        buttons=QDialogButtonBox(QDialogButtonBox.Save|QDialogButtonBox.Cancel);buttons.accepted.connect(self.accept);buttons.rejected.connect(self.reject);form.addRow(buttons)

    def values(self):
        end=None if self.end.date()==self.end.minimumDate() else self.end.date().toString("yyyy-MM-dd")
        return {"name":self.name.text().strip(),"start_date":self.start.date().toString("yyyy-MM-dd"),"end_date":end,"status":self.status.currentText(),"dose":self.dose.value() or None,"unit":self.unit.text().strip(),"frequency":self.frequency.text().strip(),"route":self.route.text().strip(),"prescriber":self.doctor.text().strip(),"notes":self.notes.toPlainText().strip(),"needs_confirmation":self.status.currentText()=="待确认" or self.dose.value()==0}


class SupportAuthorPage(QWidget):
    def __init__(self, payment_qrs, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self)
        root.setContentsMargins(32, 28, 32, 28)
        card = QFrame()
        card.setObjectName("Card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(32, 28, 32, 30)
        title = QLabel("支持身高小助理")
        title.setObjectName("PageTitle")
        title.setAlignment(Qt.AlignCenter)
        desc = QLabel("如果这个小工具对你有帮助，可以扫码支持作者继续维护和改进。\n打赏完全自愿，不影响任何功能的使用。")
        desc.setObjectName("Muted")
        desc.setWordWrap(True)
        desc.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(title)
        card_layout.addWidget(desc)
        card_layout.addSpacing(8)
        qr_row = QHBoxLayout()
        qr_row.setSpacing(24)
        qr_row.addStretch()
        for payment_name, qr_path in payment_qrs:
            payment_box = QVBoxLayout()
            payment_title = QLabel(payment_name)
            payment_title.setObjectName("CardTitle")
            payment_title.setAlignment(Qt.AlignCenter)
            qr = QLabel()
            qr.setObjectName("QrPanel")
            qr.setAlignment(Qt.AlignCenter)
            qr.setFixedSize(300, 350)
            pixmap = QPixmap(str(qr_path))
            qr.setPixmap(pixmap.scaled(270, 320, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            payment_box.addWidget(payment_title)
            payment_box.addWidget(qr)
            qr_row.addLayout(payment_box)
        qr_row.addStretch()
        card_layout.addLayout(qr_row)
        root.addWidget(card)
        root.addStretch()


class SettingsPage(QWidget):
    changed = Signal(int)

    def __init__(self, db, current_child):
        super().__init__()
        self.db = db
        self.current_child = current_child
        self.editing_id = None
        self.editing_archived = False

        root = QVBoxLayout(self)
        self.tabs = QTabWidget()
        root.addWidget(self.tabs)

        profile = QWidget()
        pl = QHBoxLayout(profile)
        left = QVBoxLayout()
        left.addWidget(QLabel("孩子列表"))
        self.child_filter = QComboBox()
        self.child_filter.addItems(["正在使用", "已归档", "全部孩子"])
        self.child_filter.currentIndexChanged.connect(self.filter_changed)
        left.addWidget(self.child_filter)
        self.child_list = QListWidget()
        self.child_list.setMinimumWidth(230)
        self.child_list.setMaximumWidth(270)
        self.child_list.currentRowChanged.connect(self.load_selected)
        left.addWidget(self.child_list, 1)
        list_tip = QLabel("归档只会停止显示，所有资料、身高记录和用药信息都会保留。")
        list_tip.setObjectName("Muted")
        list_tip.setWordWrap(True)
        list_tip.setMaximumWidth(260)
        left.addWidget(list_tip)
        pl.addLayout(left)

        right = QVBoxLayout()
        status_row = QHBoxLayout()
        self.profile_status = QLabel("新资料")
        self.profile_status.setObjectName("Badge")
        self.profile_summary = QLabel()
        self.profile_summary.setObjectName("Muted")
        status_row.addWidget(self.profile_status)
        status_row.addWidget(self.profile_summary)
        status_row.addStretch()
        right.addLayout(status_row)

        self.form_box = QGroupBox("基本资料和关注需求")
        form = QFormLayout(self.form_box)
        self.name = QLineEdit()
        self.nickname = QLineEdit()
        self.gender = QComboBox()
        self.gender.addItems(["男", "女"])
        self.birth = QDateEdit(QDate.currentDate())
        self.birth.setCalendarPopup(True)
        self.birth.setDisplayFormat("yyyy-MM-dd")
        self.father = QDoubleSpinBox()
        self.father.setRange(0, 250)
        self.father.setSpecialValueText("未填写")
        self.father.setSuffix(" cm")
        self.mother = QDoubleSpinBox()
        self.mother.setRange(0, 250)
        self.mother.setSpecialValueText("未填写")
        self.mother.setSuffix(" cm")
        self.focus = QTextEdit()
        self.focus.setPlaceholderText("例如：关注年增长速度、用药阶段、复查提醒")
        self.focus.setMaximumHeight(90)
        self.form_widgets = [self.name, self.nickname, self.gender, self.birth, self.father, self.mother, self.focus]
        for label, widget in (("姓名", self.name), ("昵称", self.nickname), ("性别", self.gender),
                              ("出生日期", self.birth), ("父亲身高", self.father),
                              ("母亲身高", self.mother), ("关注需求", self.focus)):
            form.addRow(label, widget)
        right.addWidget(self.form_box)

        buttons = QHBoxLayout()
        self.new_button = QPushButton("新增孩子")
        self.new_button.clicked.connect(self.new_child)
        self.save_button = QPushButton("保存资料")
        self.save_button.setObjectName("Primary")
        self.save_button.clicked.connect(self.save_child)
        self.archive_button = QPushButton("停止显示（归档）")
        self.archive_button.setObjectName("Danger")
        self.archive_button.clicked.connect(self.archive_child)
        self.restore_button = QPushButton("恢复使用")
        self.restore_button.setObjectName("Primary")
        self.restore_button.clicked.connect(self.restore_child)
        buttons.addWidget(self.new_button)
        buttons.addStretch()
        buttons.addWidget(self.archive_button)
        buttons.addWidget(self.restore_button)
        buttons.addWidget(self.save_button)
        right.addLayout(buttons)
        right.addStretch()
        pl.addLayout(right, 1)
        self.tabs.addTab(profile, "孩子与需求")

        meds = QWidget()
        ml = QVBoxLayout(meds)
        top = QHBoxLayout()
        self.add_med_button = QPushButton("＋ 添加用药阶段")
        self.add_med_button.setObjectName("Primary")
        self.add_med_button.clicked.connect(self.add_period)
        top.addWidget(self.add_med_button)
        self.med_hint = QLabel()
        self.med_hint.setObjectName("Muted")
        top.addWidget(self.med_hint)
        top.addStretch()
        ml.addLayout(top)
        self.med_table = QTableWidget(0, 7)
        self.med_table.setHorizontalHeaderLabels(["药品", "开始", "结束", "状态", "剂量", "频次", "说明"])
        self.med_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        ml.addWidget(self.med_table)
        self.tabs.addTab(meds, "用药管理")

        payment_qrs = []
        qr_definitions = (
            ("微信支付", "wechat-reward-qr"),
            ("支付宝", "alipay-reward-qr"),
        )
        for payment_name, stem in qr_definitions:
            candidates = [
                app_root() / "support" / f"{stem}.jpg",
                app_root() / "support" / f"{stem}.png",
                bundled_path("assets", f"{stem}.jpg"),
                bundled_path("assets", f"{stem}.png"),
            ]
            qr_path = next((path for path in candidates if path.exists()), None)
            if qr_path:
                payment_qrs.append((payment_name, qr_path))
        if payment_qrs:
            self.tabs.addTab(SupportAuthorPage(payment_qrs), "支持作者")

    def visible_children(self):
        children = self.db.children(include_archived=True)
        if self.child_filter.currentIndex() == 0:
            return [child for child in children if not child["archived"]]
        if self.child_filter.currentIndex() == 1:
            return [child for child in children if child["archived"]]
        return children

    def filter_changed(self):
        self.refresh()

    def refresh(self, selected_id=None):
        children = self.visible_children()
        self.child_list.blockSignals(True)
        self.child_list.clear()
        target = -1
        for i, child in enumerate(children):
            label = f'{child["name"]}  ·  已归档' if child["archived"] else child["name"]
            self.child_list.addItem(label)
            self.child_list.item(i).setData(Qt.UserRole, child["id"])
            if child["id"] == selected_id:
                target = i
        self.child_list.blockSignals(False)
        if children:
            target = target if target >= 0 else 0
            self.child_list.setCurrentRow(target)
            self.load_selected(target)
        else:
            self.clear_profile_for_empty_filter()
        self.refresh_meds()

    def clear_profile_for_empty_filter(self):
        self.editing_id = None
        self.editing_archived = False
        self.name.clear()
        self.nickname.clear()
        self.gender.setCurrentIndex(0)
        self.birth.setDate(QDate.currentDate())
        self.father.setValue(0)
        self.mother.setValue(0)
        self.focus.clear()
        archived_view = self.child_filter.currentIndex() == 1
        self.profile_status.setText("暂无已归档孩子" if archived_view else "新资料")
        self.profile_status.setObjectName("ArchiveBadge" if archived_view else "Badge")
        self.profile_status.style().unpolish(self.profile_status)
        self.profile_status.style().polish(self.profile_status)
        self.profile_summary.setText("" if not archived_view else "归档后的孩子会显示在这里，并可随时恢复")
        self.set_profile_mode("new")

    def load_selected(self, row):
        if row < 0 or not self.child_list.item(row):
            return
        child = self.db.child(self.child_list.item(row).data(Qt.UserRole))
        if not child:
            return
        self.editing_id = child["id"]
        self.editing_archived = bool(child["archived"])
        self.name.setText(child["name"])
        self.nickname.setText(child["nickname"])
        self.gender.setCurrentText(child["gender"])
        self.birth.setDate(QDate.fromString(child["birth_date"], "yyyy-MM-dd"))
        self.father.setValue(child["father_height"] or 0)
        self.mother.setValue(child["mother_height"] or 0)
        self.focus.setPlainText(child["focus"])
        summary = self.db.child_summary(child["id"])
        self.profile_summary.setText(
            f'{summary["measurements"]} 条测量记录 · {summary["medications"]} 种药品 · {summary["periods"]} 个用药阶段'
        )
        self.profile_status.setText("已归档" if self.editing_archived else "正在使用")
        self.profile_status.setObjectName("ArchiveBadge" if self.editing_archived else "Badge")
        self.profile_status.style().unpolish(self.profile_status)
        self.profile_status.style().polish(self.profile_status)
        self.set_profile_mode("archived" if self.editing_archived else "active")
        self.refresh_meds()

    def set_profile_mode(self, mode):
        editable = mode != "archived"
        for widget in self.form_widgets:
            widget.setEnabled(editable)
        self.save_button.setVisible(mode in ("active", "new"))
        self.archive_button.setVisible(mode == "active")
        self.restore_button.setVisible(mode == "archived")
        self.add_med_button.setEnabled(mode == "active")
        if mode == "archived":
            self.med_hint.setText("已归档：用药资料只读，恢复后可继续编辑")
        elif mode == "new":
            self.med_hint.setText("请先保存孩子资料，再添加用药阶段")
        else:
            self.med_hint.setText("")

    def new_child(self):
        if self.child_filter.currentIndex() != 0:
            self.child_filter.blockSignals(True)
            self.child_filter.setCurrentIndex(0)
            self.child_filter.blockSignals(False)
        self.child_list.clearSelection()
        self.editing_id = None
        self.editing_archived = False
        self.name.clear()
        self.nickname.clear()
        self.gender.setCurrentIndex(0)
        self.birth.setDate(QDate.currentDate())
        self.father.setValue(0)
        self.mother.setValue(0)
        self.focus.clear()
        self.profile_status.setText("新资料")
        self.profile_status.setObjectName("Badge")
        self.profile_status.style().unpolish(self.profile_status)
        self.profile_status.style().polish(self.profile_status)
        self.profile_summary.setText("填写后点击“保存资料”")
        self.set_profile_mode("new")
        self.refresh_meds()
        self.name.setFocus()

    def save_child(self):
        if self.editing_archived:
            return
        if not self.name.text().strip():
            QMessageBox.warning(self, APP_NAME, "请填写孩子姓名。")
            return
        values = {"name": self.name.text().strip(), "nickname": self.nickname.text().strip(),
                  "gender": self.gender.currentText(), "birth_date": self.birth.date().toString("yyyy-MM-dd"),
                  "father_height": self.father.value() or None, "mother_height": self.mother.value() or None,
                  "focus": self.focus.toPlainText().strip()}
        child_id = self.db.save_child(values, self.editing_id)
        self.editing_id = child_id
        self.changed.emit(child_id)
        QMessageBox.information(self, APP_NAME, "基本资料和关注需求已保存。")

    def archive_child(self):
        if not self.editing_id or self.editing_archived:
            return
        child = self.db.child(self.editing_id)
        summary = self.db.child_summary(self.editing_id)
        message = (f'要停止显示“{child["name"]}”吗？\n\n'
                   f'将保留：{summary["measurements"]} 条测量记录、{summary["medications"]} 种药品、'
                   f'{summary["periods"]} 个用药阶段。\n\n'
                   '之后可在“已归档”列表中查看并恢复，任何原始数据都不会删除。')
        if QMessageBox.question(self, APP_NAME, message) != QMessageBox.Yes:
            return
        child_id = self.editing_id
        self.db.archive_child(child_id)
        self.child_filter.setCurrentIndex(1)
        self.refresh(child_id)
        self.changed.emit(0)

    def restore_child(self):
        if not self.editing_id or not self.editing_archived:
            return
        child = self.db.child(self.editing_id)
        if QMessageBox.question(
            self, APP_NAME,
            f'恢复“{child["name"]}”继续使用吗？\n\n恢复后会重新出现在顶部孩子选择框、首页、历史记录和图表中。'
        ) != QMessageBox.Yes:
            return
        child_id = self.editing_id
        self.db.restore_child(child_id)
        self.child_filter.setCurrentIndex(0)
        self.refresh(child_id)
        self.changed.emit(child_id)

    def add_period(self):
        child = self.db.child(self.editing_id) if self.editing_id else None
        if not child or child["archived"]:
            QMessageBox.warning(self, APP_NAME, "请先选择一个正在使用的孩子。")
            return
        dialog = MedicationDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return
        values = dialog.values()
        if not values["name"]:
            QMessageBox.warning(self, APP_NAME, "请输入药品名称。")
            return
        med_id = self.db.get_or_create_medication(
            child["id"], values.pop("name"), default_unit=values["unit"],
            route=values["route"], prescriber=values["prescriber"]
        )
        self.db.add_medication_period(med_id, **values)
        self.refresh_meds()
        self.changed.emit(child["id"])

    def refresh_meds(self):
        child_id = self.editing_id
        rows = self.db.medication_periods(child_id) if child_id else []
        self.med_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            dose = "待确认" if row["dose"] is None else f'{row["dose"]} {row["unit"]}'
            vals = (row["medication_name"], row["start_date"], row["end_date"] or "至今", row["status"],
                    dose, row["frequency"], row["notes"])
            for c, val in enumerate(vals):
                self.med_table.setItem(r, c, item(val, Qt.AlignLeft | Qt.AlignVCenter if c == 6 else Qt.AlignCenter))


class BackupPage(QWidget):
    changed=Signal()
    def __init__(self,db):
        super().__init__();self.db=db;root=QVBoxLayout(self)
        box=QGroupBox("本地数据安全");layout=QGridLayout(box)
        entries=[("选择位置并备份","由你决定保存到 U 盘、移动硬盘、网盘同步目录或电脑中的任意文件夹。",self.backup),("从备份恢复","选择以前生成的数据库备份；恢复前会先让你指定当前数据的安全副本位置。",self.restore),("打开程序数据位置","查看当前数据库和导出文件。普通使用无需手动修改其中内容。",self.open_folder)]
        for r,(title,desc,func) in enumerate(entries):layout.addWidget(QLabel(f"<b>{title}</b><br><span style='color:#71838B'>{desc}</span>"),r,0);button=QPushButton(title);button.clicked.connect(func);layout.addWidget(button,r,1)
        root.addWidget(box)
        tip=QFrame();tip.setObjectName("InfoPanel");tip_layout=QVBoxLayout(tip);tip_layout.addWidget(QLabel("备份建议"));advice=QLabel("重要数据建议至少保留两份，例如：电脑一份 + U 盘或私人网盘一份。备份文件可以直接复制，但请不要用 Excel 打开或修改。") ;advice.setWordWrap(True);advice.setObjectName("Muted");tip_layout.addWidget(advice);root.addWidget(tip)
        self.info=QLabel();self.info.setObjectName("Muted");self.info.setWordWrap(True);root.addWidget(self.info);root.addStretch();self.refresh()

    def refresh(self):
        last=self.db.setting("last_manual_backup",{})
        last_text=last.get("path","尚未进行手动备份") if isinstance(last,dict) else str(last)
        self.info.setText(f"当前数据库：{self.db.db_path}\n最近一次手动备份：{last_text}")

    def backup(self):
        stamp=datetime.now().strftime("%Y%m%d-%H%M%S");default=Path.home()/"Documents"/f"身高小助理备份-{stamp}.db"
        path=QFileDialog.getSaveFileName(self,"选择备份保存位置",str(default),"身高小助理备份 (*.db)")[0]
        if not path:return
        if not path.lower().endswith(".db"):path += ".db"
        try:self.db.backup_to(path)
        except (OSError,ValueError) as exc:QMessageBox.warning(self,APP_NAME,f"备份失败：{exc}");return
        self.db.set_setting("last_manual_backup",{"path":path,"time":datetime.now().isoformat(timespec="seconds")});self.refresh();QMessageBox.information(self,APP_NAME,f"备份完成。请妥善保存：\n{path}")

    def restore(self):
        source=QFileDialog.getOpenFileName(self,"选择要恢复的备份",str(Path.home()/"Documents"),"身高小助理备份 (*.db);;SQLite 数据库 (*.sqlite *.sqlite3)")[0]
        if not source:return
        valid,message=self.db.validate_backup(source)
        if not valid:QMessageBox.warning(self,APP_NAME,message);return
        stamp=datetime.now().strftime("%Y%m%d-%H%M%S");default=Path(source).parent/f"恢复前的当前数据-{stamp}.db"
        safety=QFileDialog.getSaveFileName(self,"先保存当前数据的安全副本",str(default),"身高小助理备份 (*.db)")[0]
        if not safety:return
        if not safety.lower().endswith(".db"):safety += ".db"
        if QMessageBox.question(self,APP_NAME,"即将用所选备份替换当前数据。当前数据会先保存到你刚指定的位置，是否继续？")!=QMessageBox.Yes:return
        self.db.backup_to(safety);self.db.restore_from(source);self.db.set_setting("last_manual_backup",{"path":safety,"time":datetime.now().isoformat(timespec="seconds")});self.refresh();self.changed.emit();QMessageBox.information(self,APP_NAME,"恢复完成，孩子、记录和图表已经重新载入。")

    def open_folder(self):
        try: os.startfile(self.db.data_dir.parent)
        except OSError: QMessageBox.information(self,APP_NAME,str(self.db.data_dir.parent))


class MainWindow(QMainWindow):
    def __init__(self,db):
        super().__init__();self.db=db;self.setWindowTitle(APP_NAME);self.resize(1280,820);self.setMinimumSize(1050,700)
        root=BackgroundWidget(bundled_path("assets","client-background-v2.png"));root.setObjectName("AppRoot");self.setCentralWidget(root);outer=QHBoxLayout(root);outer.setContentsMargins(0,0,0,0);outer.setSpacing(0)
        sidebar=QFrame();sidebar.setObjectName("Sidebar");sidebar.setFixedWidth(220);side=QVBoxLayout(sidebar);side.setContentsMargins(14,14,14,18)
        brand_row=QHBoxLayout();brand_icon=QLabel();brand_pix=QPixmap(str(bundled_path("assets","client-icon-v2.png")));brand_icon.setPixmap(brand_pix.scaled(54,54,Qt.KeepAspectRatio,Qt.SmoothTransformation));brand_icon.setFixedSize(58,58);brand_row.addWidget(brand_icon)
        brand_text=QVBoxLayout();brand=QLabel("身高小助理");brand.setObjectName("Brand");brand_text.addWidget(brand);sub=QLabel("陪孩子稳稳长大");sub.setObjectName("BrandSub");brand_text.addWidget(sub);brand_row.addLayout(brand_text);side.addLayout(brand_row);side.addSpacing(18)
        self.nav=[]
        labels=[("●  首页概览",0),("＋  记录身高",1),("≡  历史记录",2),("↗  深度图表",3),("▣  备份与恢复",4),("⚙  设置",5)]
        for label,index in labels:
            b=QPushButton(label);b.setObjectName("NavButton");b.setCheckable(True);b.clicked.connect(lambda checked,i=index:self.switch_page(i));side.addWidget(b);self.nav.append(b)
        side.addStretch();privacy=QLabel("数据仅保存在本机\n不会自动上传云端");privacy.setObjectName("BrandSub");privacy.setAlignment(Qt.AlignCenter);side.addWidget(privacy);outer.addWidget(sidebar)
        content=QVBoxLayout();content.setContentsMargins(0,0,0,0);content.setSpacing(0);outer.addLayout(content,1)
        top=QFrame();top.setObjectName("TopBar");tl=QHBoxLayout(top);tl.setContentsMargins(26,14,26,14);titles=QVBoxLayout();self.page_title=QLabel();self.page_title.setObjectName("PageTitle");self.page_hint=QLabel();self.page_hint.setObjectName("PageHint");titles.addWidget(self.page_title);titles.addWidget(self.page_hint);tl.addLayout(titles);tl.addStretch();tl.addWidget(QLabel("当前孩子"));self.child_combo=QComboBox();self.child_combo.setMinimumWidth(160);self.child_combo.currentIndexChanged.connect(self.child_changed);tl.addWidget(self.child_combo);content.addWidget(top)
        body=QWidget();body.setAttribute(Qt.WA_TranslucentBackground);bl=QVBoxLayout(body);bl.setContentsMargins(24,20,24,24);self.stack=QStackedWidget();bl.addWidget(self.stack);content.addWidget(body,1)
        self.pages=[HomePage(db),AddPage(db,self.current_child),HistoryPage(db,self.current_child),ChartsPage(db,self.current_child),BackupPage(db),SettingsPage(db,self.current_child)]
        for page in self.pages:self.stack.addWidget(page)
        self.pages[1].saved.connect(self.refresh_all);self.pages[2].changed.connect(self.refresh_all);self.pages[4].changed.connect(self.reload_children);self.pages[5].changed.connect(self.reload_children)
        self.reload_children();self.switch_page(0)

    def current_child(self):
        cid=self.child_combo.currentData();return self.db.child(cid) if cid else None

    def reload_children(self,select_id=None):
        if isinstance(select_id,bool):select_id=None
        current=select_id or self.child_combo.currentData();children=self.db.children();self.child_combo.blockSignals(True);self.child_combo.clear();target=0
        for i,child in enumerate(children):self.child_combo.addItem(child["name"],child["id"]);target=i if child["id"]==current else target
        self.child_combo.blockSignals(False)
        if children:self.child_combo.setCurrentIndex(target)
        self.pages[5].refresh(self.child_combo.currentData());self.refresh_all()
        if not children:self.switch_page(5)

    def child_changed(self):
        self.pages[5].refresh(self.child_combo.currentData());self.refresh_all()

    def refresh_all(self):
        child=self.current_child();self.pages[0].refresh(child);self.pages[2].refresh(child);self.pages[3].refresh(child);self.pages[4].refresh()

    def switch_page(self,index):
        titles=[("首页概览","最近一次记录和整体趋势"),("记录身高","新增一次测量，可同时记录多种药品"),("历史记录","搜索、正序或倒序查看并导出全部记录"),("深度图表","悬浮查看数据点，分析身高、速度、体重、BMI 与用药阶段"),("备份与恢复","备份保存位置完全由你决定"),("设置","管理任意数量的孩子、需求和用药阶段")]
        self.stack.setCurrentIndex(index);self.page_title.setText(titles[index][0]);self.page_hint.setText(titles[index][1])
        for i,b in enumerate(self.nav):b.setChecked(i==index)
        if index==1:self.pages[1].load_active_meds()
        if index==5:self.pages[5].refresh(self.child_combo.currentData())


def main():
    parser=argparse.ArgumentParser();parser.add_argument("--data-dir");parser.add_argument("--no-gui",action="store_true");args=parser.parse_args()
    db=HeightDatabase(args.data_dir or default_data_dir())
    if args.no_gui:
        print(json.dumps({"database":str(db.db_path),"children":len(db.children())},ensure_ascii=False));return 0
    app=QApplication(sys.argv);app.setApplicationName(APP_NAME);configure_chinese_font(app)
    arrow_path=bundled_path("assets","combo-arrow.svg").as_posix()
    calendar_path=bundled_path("assets","calendar.svg").as_posix()
    spin_up_path=bundled_path("assets","spin-up.svg").as_posix()
    spin_down_path=bundled_path("assets","spin-down.svg").as_posix()
    style=APP_STYLE.replace("__COMBO_ARROW__",arrow_path).replace("__CALENDAR_ICON__",calendar_path).replace("__SPIN_UP__",spin_up_path).replace("__SPIN_DOWN__",spin_down_path)
    app.setStyleSheet(style)
    icon_path=bundled_path("assets","client-icon-v2.png")
    if icon_path.exists():app.setWindowIcon(QIcon(str(icon_path)))
    window=MainWindow(db);window.show();return app.exec()


if __name__=="__main__":raise SystemExit(main())
